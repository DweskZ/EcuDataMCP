import csv
import io
import json
import logging
import re
import tarfile
import zipfile
import zlib
from typing import Any

import httpx

from helpers.logging import MAIN_LOGGER_NAME
from helpers.safe_download import safe_stream
from helpers.tls import should_retry_insecure
from helpers.user_agent import USER_AGENT

logger = logging.getLogger(MAIN_LOGGER_NAME)

MAX_DOWNLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
_TIMEOUT = 30.0

_GEOM_COLUMN_NAMES = {"geom", "geometry", "the_geom", "wkt", "wkt_geom", "shape"}
_WKT_PREFIXES = (
    "POINT",
    "LINESTRING",
    "POLYGON",
    "MULTIPOINT",
    "MULTILINESTRING",
    "MULTIPOLYGON",
    "GEOMETRYCOLLECTION",
)

# "7.760,2" (thousands dot, decimal comma) or plain "168,15".
_EU_DECIMAL_RE = re.compile(r"^-?(?:\d{1,3}(?:\.\d{3})+|\d+),\d+$")


def _looks_like_wkt(value: str) -> bool:
    return value.strip().upper().startswith(_WKT_PREFIXES)


def strip_geometry_columns(
    headers: list[str], rows: list[list[str]]
) -> tuple[list[str], list[list[str]], list[str]]:
    """Drop columns that look like WKT/geometry data, by column name or content.

    A single polygon/multipolygon cell can be tens of KB of coordinates,
    which drowns out the rest of a preview. Returns (headers, rows, dropped_names).
    """
    if not headers:
        return headers, rows, []

    drop_idx: list[int] = []
    for i, name in enumerate(headers):
        if name.strip().lower() in _GEOM_COLUMN_NAMES:
            drop_idx.append(i)
            continue
        sample = [row[i] for row in rows[:5] if i < len(row) and row[i]]
        if sample and all(_looks_like_wkt(v) for v in sample):
            drop_idx.append(i)

    if not drop_idx:
        return headers, rows, []

    drop_set = set(drop_idx)
    dropped_names = [headers[i] for i in drop_idx]
    new_headers = [h for i, h in enumerate(headers) if i not in drop_set]
    new_rows = [
        [cell for i, cell in enumerate(row) if i not in drop_set] for row in rows
    ]
    return new_headers, new_rows, dropped_names


def _convert_eu_decimal(value: str) -> str:
    v = value.strip()
    sign = "-" if v.startswith("-") else ""
    v = v.lstrip("-").replace(".", "").replace(",", ".")
    return sign + v


def normalize_eu_decimal_columns(
    headers: list[str], rows: list[list[str]]
) -> tuple[list[list[str]], list[str]]:
    """Convert columns formatted as European decimals (7.760,2 -> 7760.2).

    Ecuadorian government CSVs commonly use dot-thousands/comma-decimal
    notation. Left as-is, those values sort and parse wrong downstream.
    A column only qualifies if every sampled value matches the pattern, so
    genuinely ambiguous columns are left untouched. Returns (rows, converted_names).
    """
    if not headers or not rows:
        return rows, []

    idxs: list[int] = []
    for i in range(len(headers)):
        sample = [row[i] for row in rows[:5] if i < len(row) and row[i]]
        if sample and all(_EU_DECIMAL_RE.match(v.strip()) for v in sample):
            idxs.append(i)

    if not idxs:
        return rows, []

    idx_set = set(idxs)
    new_rows = [
        [
            _convert_eu_decimal(cell) if i in idx_set and cell else cell
            for i, cell in enumerate(row)
        ]
        for row in rows
    ]
    return new_rows, [headers[i] for i in idxs]


async def _download(session: httpx.AsyncClient, url: str) -> tuple[bytes, bool]:
    truncated = False
    # `url` comes from CKAN resource metadata -- external, not first-party --
    # so every hop (including redirects) must be checked against the SSRF
    # guard rather than trusting httpx's own follow_redirects.
    async with safe_stream(session, url, timeout=_TIMEOUT) as resp:
        resp.raise_for_status()

        content_length = resp.headers.get("content-length")
        if content_length and int(content_length) > MAX_DOWNLOAD_BYTES:
            truncated = True

        chunks: list[bytes] = []
        total = 0
        async for chunk in resp.aiter_bytes(chunk_size=64 * 1024):
            chunks.append(chunk)
            total += len(chunk)
            if total > MAX_DOWNLOAD_BYTES:
                truncated = True
                break

    return b"".join(chunks), truncated


async def download_bytes(
    url: str, session: httpx.AsyncClient | None = None
) -> tuple[bytes, bool]:
    """Download up to MAX_DOWNLOAD_BYTES with TLS fallback for portal hosts."""
    own = session is None
    if own:
        # follow_redirects is deliberately not set here: safe_stream() always
        # makes its own per-request follow_redirects=False and handles
        # redirects itself (validating each hop), so a client-level default
        # would be misleading rather than wrong.
        session = httpx.AsyncClient(headers={"User-Agent": USER_AGENT})
    assert session is not None
    try:
        logger.debug("Downloading from %s (max %d bytes)", url, MAX_DOWNLOAD_BYTES)
        try:
            return await _download(session, url)
        except httpx.ConnectError as exc:
            if not should_retry_insecure(exc, url):
                raise
            logger.warning(
                "TLS verification failed for %s (portal cert expired); "
                "retrying without verification",
                url,
            )
            async with httpx.AsyncClient(
                headers={"User-Agent": USER_AGENT},
                verify=False,
            ) as insecure_session:
                return await _download(insecure_session, url)
    finally:
        if own:
            await session.aclose()


async def sniff_content_type(
    url: str, session: httpx.AsyncClient | None = None
) -> str | None:
    """Best-effort Content-Type sniff for a resource with no recognizable
    URL extension and no useful declared CKAN format.

    Opens the same safe_stream() a real download would use but returns as
    soon as headers arrive, without reading the body -- costs a connection,
    not a download. Best-effort: any failure just means no hint, callers
    fall back to their existing "unsupported format" handling.
    """
    own = session is None
    if own:
        session = httpx.AsyncClient(headers={"User-Agent": USER_AGENT})
    assert session is not None
    try:
        async with safe_stream(session, url, timeout=_TIMEOUT) as resp:
            return resp.headers.get("content-type")
    except httpx.HTTPError:
        return None
    finally:
        if own:
            await session.aclose()


def _decode_text(raw: bytes) -> str:
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return raw.decode(encoding)
        except (UnicodeDecodeError, ValueError):
            continue
    return raw.decode("utf-8", errors="replace")


def _parse_csv_bytes(raw: bytes, max_rows: int, truncated: bool = False) -> dict[str, Any]:
    text = _decode_text(raw)

    sample = text[:2000]
    delimiter = ","
    for candidate in (";", "\t", "|"):
        if sample.count(candidate) > sample.count(delimiter):
            delimiter = candidate

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows_read: list[list[str]] = []
    try:
        for row in reader:
            rows_read.append(row)
            if len(rows_read) > max_rows + 1:
                truncated = True
                break
    except csv.Error as e:
        # Found for real: a .zip's picked "first tabular-looking member" can
        # still turn out to be a binary file misidentified by its extension,
        # or a genuinely malformed export. Keep whatever rows parsed cleanly
        # before the bad one rather than losing the whole preview to it.
        truncated = True
        if not rows_read:
            raise ValueError(
                "El contenido no se pudo parsear como CSV (formato inválido "
                f"o no es realmente texto tabular): {e}"
            ) from e

    if not rows_read:
        return {"headers": [], "rows": [], "total_rows_in_preview": 0, "truncated": truncated}

    headers = rows_read[0]
    data_rows = rows_read[1 : max_rows + 1]

    headers, data_rows, dropped_geom = strip_geometry_columns(headers, data_rows)
    data_rows, converted_decimals = normalize_eu_decimal_columns(headers, data_rows)

    result: dict[str, Any] = {
        "headers": headers,
        "rows": data_rows,
        "total_rows_in_preview": len(data_rows),
        "truncated": truncated,
    }
    if dropped_geom:
        result["dropped_columns"] = dropped_geom
    if converted_decimals:
        result["converted_decimal_columns"] = converted_decimals
    return result


async def preview_csv(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """
    Download a CSV file and return the first N rows parsed.

    Returns:
        dict with 'headers', 'rows', 'total_rows_in_preview', 'truncated'
    """
    raw, truncated = await download_bytes(url, session=session)
    result = _parse_csv_bytes(raw, max_rows, truncated=truncated)
    result["format"] = "csv"
    return result


async def preview_json(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview JSON / GeoJSON as a tabular sample when possible."""
    raw, truncated = await download_bytes(url, session=session)
    text = _decode_text(raw)
    data = json.loads(text)

    if isinstance(data, dict) and "features" in data and isinstance(data["features"], list):
        features = data["features"]
        rows: list[list[str]] = []
        headers = ["id", "geometry_type", "properties"]
        for feat in features[:max_rows]:
            props = feat.get("properties") or {}
            geom = feat.get("geometry") or {}
            rows.append(
                [
                    str(feat.get("id", "")),
                    str(geom.get("type", "")),
                    json.dumps(props, ensure_ascii=False)[:200],
                ]
            )
        return {
            "headers": headers,
            "rows": rows,
            "total_rows_in_preview": len(rows),
            "truncated": truncated or len(features) > max_rows,
            "format": "geojson",
            "total_records": len(features),
        }

    if isinstance(data, list):
        if not data:
            return {
                "headers": [],
                "rows": [],
                "total_rows_in_preview": 0,
                "truncated": truncated,
                "format": "json",
            }
        if all(isinstance(item, dict) for item in data[:max_rows]):
            keys: list[str] = []
            for item in data[:max_rows]:
                for key in item:
                    if key not in keys:
                        keys.append(key)
            rows = [
                [str(item.get(k, ""))[:120] for k in keys] for item in data[:max_rows]
            ]
            keys, rows, dropped_geom = strip_geometry_columns(keys, rows)
            result: dict[str, Any] = {
                "headers": keys,
                "rows": rows,
                "total_rows_in_preview": len(rows),
                "truncated": truncated or len(data) > max_rows,
                "format": "json",
                "total_records": len(data),
            }
            if dropped_geom:
                result["dropped_columns"] = dropped_geom
            return result
        preview_items = data[:max_rows]
        return {
            "headers": ["value"],
            "rows": [[json.dumps(item, ensure_ascii=False)[:200]] for item in preview_items],
            "total_rows_in_preview": len(preview_items),
            "truncated": truncated or len(data) > max_rows,
            "format": "json",
            "total_records": len(data),
        }

    if isinstance(data, dict):
        headers = ["key", "value"]
        items = list(data.items())
        rows = [[str(k), str(v)[:200]] for k, v in items[:max_rows]]
        return {
            "headers": headers,
            "rows": rows,
            "total_rows_in_preview": len(rows),
            "truncated": truncated or len(items) > max_rows,
            "format": "json",
            "total_records": len(items),
        }

    return {
        "headers": ["value"],
        "rows": [[str(data)[:500]]],
        "total_rows_in_preview": 1,
        "truncated": truncated,
        "format": "json",
    }


async def preview_xlsx(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview the first sheet of an Excel workbook."""
    from openpyxl import load_workbook

    raw, truncated = await download_bytes(url, session=session)
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return {
                "headers": [],
                "rows": [],
                "total_rows_in_preview": 0,
                "truncated": truncated,
                "format": "xlsx",
            }
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {
                "headers": [],
                "rows": [],
                "total_rows_in_preview": 0,
                "truncated": truncated,
                "format": "xlsx",
            }
        headers = [str(c) if c is not None else "" for c in header_row]
        data_rows: list[list[str]] = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                truncated = True
                break
            data_rows.append([str(c) if c is not None else "" for c in row])
        return {
            "headers": headers,
            "rows": data_rows,
            "total_rows_in_preview": len(data_rows),
            "truncated": truncated,
            "format": "xlsx",
            "sheet": ws.title,
        }
    finally:
        wb.close()


async def preview_xls(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview the first sheet of a legacy Excel (.xls) workbook."""
    import xlrd

    raw, truncated = await download_bytes(url, session=session)
    wb = xlrd.open_workbook(file_contents=raw)
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        return {
            "headers": [],
            "rows": [],
            "total_rows_in_preview": 0,
            "truncated": truncated,
            "format": "xls",
        }

    def cell_str(value: Any) -> str:
        return "" if value is None else str(value)

    headers = [cell_str(v) for v in ws.row_values(0)]
    data_rows: list[list[str]] = []
    for i in range(1, ws.nrows):
        if len(data_rows) >= max_rows:
            truncated = True
            break
        data_rows.append([cell_str(v) for v in ws.row_values(i)])

    return {
        "headers": headers,
        "rows": data_rows,
        "total_rows_in_preview": len(data_rows),
        "truncated": truncated,
        "format": "xls",
        "sheet": ws.name,
    }


async def preview_ods(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview the first sheet of an OpenDocument Spreadsheet (.ods)."""
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.teletype import extractText

    raw, truncated = await download_bytes(url, session=session)
    doc = load(io.BytesIO(raw))
    tables = doc.spreadsheet.getElementsByType(Table)
    empty = {
        "headers": [],
        "rows": [],
        "total_rows_in_preview": 0,
        "truncated": truncated,
        "format": "ods",
    }
    if not tables:
        return empty
    sheet = tables[0]

    def row_values(row: TableRow) -> list[str]:
        values: list[str] = []
        for cell in row.getElementsByType(TableCell):
            # ODS pads trailing empty columns with huge repeat counts (e.g.
            # 1024) to fill the sheet's fixed grid -- cap it, real header/data
            # rows never need more than a handful of repeated cells.
            repeat = min(int(cell.getAttribute("numbercolumnsrepeated") or 1), 50)
            values.extend([extractText(cell).strip()] * repeat)
        while values and values[-1] == "":
            values.pop()
        return values

    all_rows: list[list[str]] = []
    for row in sheet.getElementsByType(TableRow):
        values = row_values(row)
        if not values:
            continue
        repeat = min(int(row.getAttribute("numberrowsrepeated") or 1), max_rows + 2)
        for _ in range(repeat):
            all_rows.append(values)
            if len(all_rows) > max_rows + 1:
                break
        if len(all_rows) > max_rows + 1:
            break

    if not all_rows:
        return empty

    headers = all_rows[0]
    data_rows = all_rows[1 : max_rows + 1]
    if len(all_rows) - 1 > max_rows:
        truncated = True

    return {
        "headers": headers,
        "rows": data_rows,
        "total_rows_in_preview": len(data_rows),
        "truncated": truncated,
        "format": "ods",
        "sheet": sheet.getAttribute("name"),
    }


_MAX_DECOMPRESSED_BYTES = 20 * 1024 * 1024  # 20 MB


def _gunzip_capped(raw: bytes, cap: int = _MAX_DECOMPRESSED_BYTES) -> tuple[bytes, bool]:
    """Decompress gzip bytes, stopping once `cap` output bytes are produced.

    Bounds memory use against a decompression bomb: a small, highly
    compressible .tar.gz (already capped at MAX_DOWNLOAD_BYTES on the wire)
    could otherwise expand to gigabytes once decompressed.
    """
    decompressor = zlib.decompressobj(16 + zlib.MAX_WBITS)
    out = bytearray()
    chunk_size = 64 * 1024
    for i in range(0, len(raw), chunk_size):
        out += decompressor.decompress(raw[i : i + chunk_size])
        if len(out) > cap:
            return bytes(out[:cap]), True
    out += decompressor.flush()
    return bytes(out[:cap]), len(out) > cap


def _pick_member(names: list[str]) -> str | None:
    """Pick the best-priority member name: .csv > .tsv > .txt, or None if no
    member looks tabular.

    Prevents a bundled readme.txt from winning over the actual data file
    just because it comes first in the archive. Deliberately does NOT fall
    back to "just take the first file" -- confirmed against a real GIS
    raster .zip on the live portal (.lyr/.tif/.tif.aux.xml, no CSV at all)
    that a naive first-file fallback force-feeds a binary file into the CSV
    parser, which fails with a confusing raw csv.Error instead of a message
    that explains there's simply no tabular content in the archive.
    """
    for ext in (".csv", ".tsv", ".txt"):
        match = next((n for n in names if n.lower().endswith(ext)), None)
        if match is not None:
            return match
    return None


async def preview_targz(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview the first CSV/TSV-like member inside a .tar.gz archive."""
    raw, truncated = await download_bytes(url, session=session)
    decompressed, capped = _gunzip_capped(raw)
    truncated = truncated or capped

    try:
        with tarfile.open(fileobj=io.BytesIO(decompressed), mode="r:") as tar:
            members = {m.name: m for m in tar.getmembers() if m.isfile()}
            if not members:
                return {
                    "headers": [],
                    "rows": [],
                    "total_rows_in_preview": 0,
                    "truncated": truncated,
                    "format": "tar_gz",
                }
            member_name = _pick_member(list(members))
            if member_name is None:
                sample = ", ".join(list(members)[:10])
                raise ValueError(
                    "Este .tar.gz no contiene ningún archivo .csv/.tsv/.txt "
                    f"identificable (archivos internos: {sample}). Puede que "
                    "no sea un dataset tabular. Usa download_resource para "
                    "bajarlo completo."
                )
            extracted = tar.extractfile(members[member_name])
            inner_raw = extracted.read(MAX_DOWNLOAD_BYTES + 1) if extracted else b""
    except tarfile.TarError as e:
        if truncated:
            # Confirmed against a real 17MB .zip on the live portal (same
            # failure mode applies here): a real archive larger than the 5MB
            # download cap gets cut off mid-stream, which reads as corrupt to
            # tarfile. We already know why, so say that instead of guessing.
            raise ValueError(
                "El archivo .tar.gz descomprimido supera el límite de 5 MB de "
                "este preview y se cortó a la mitad. Usa download_resource "
                "para bajarlo completo, o el enlace directo."
            ) from e
        raise ValueError(
            "El archivo .tar.gz no se pudo leer: está corrupto"
        ) from e

    truncated = truncated or len(inner_raw) > MAX_DOWNLOAD_BYTES
    inner_raw = inner_raw[:MAX_DOWNLOAD_BYTES]

    result = _parse_csv_bytes(inner_raw, max_rows, truncated=truncated)
    result["format"] = "tar_gz"
    result["member_name"] = member_name
    return result


async def preview_zip(
    url: str, max_rows: int = 20, session: httpx.AsyncClient | None = None
) -> dict[str, Any]:
    """Preview the first CSV/TSV-like member inside a .zip archive.

    Unlike .tar.gz, a .zip's central directory lists members without
    decompressing anything, so there is no need for a capped-decompression
    pass up front — bounding the single chosen member's `.read()` call is
    enough to keep a decompression bomb from blowing up memory.
    """
    raw, truncated = await download_bytes(url, session=session)
    if truncated:
        # A .zip's central directory lives at the end of the file, so a
        # download cut off at MAX_DOWNLOAD_BYTES is missing it entirely --
        # zipfile can't open it at all, not even partially. Confirmed
        # against a real 17MB resource on the live portal: parsing wasn't
        # just degraded, it failed outright with "File is not a zip file".
        # Skip the doomed parse attempt and say what actually happened.
        raise ValueError(
            "El archivo .zip supera el límite de 5 MB de este preview, así "
            "que se descargó incompleto y no se puede abrir (el índice del "
            ".zip vive al final del archivo). Usa download_resource para "
            "bajarlo completo, o el enlace directo."
        )

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = [n for n in zf.namelist() if not n.endswith("/")]
            if not names:
                return {
                    "headers": [],
                    "rows": [],
                    "total_rows_in_preview": 0,
                    "truncated": truncated,
                    "format": "zip",
                }
            member_name = _pick_member(names)
            if member_name is None:
                sample = ", ".join(names[:10])
                raise ValueError(
                    "Este .zip no contiene ningún archivo .csv/.tsv/.txt "
                    f"identificable (archivos internos: {sample}). Puede que "
                    "no sea un dataset tabular (ej. un paquete GIS/raster). "
                    "Usa download_resource para bajarlo completo."
                )
            with zf.open(member_name) as f:
                inner_raw = f.read(MAX_DOWNLOAD_BYTES + 1)
    except zipfile.BadZipFile as e:
        raise ValueError(
            "El archivo .zip no se pudo leer: está corrupto"
        ) from e

    truncated = truncated or len(inner_raw) > MAX_DOWNLOAD_BYTES
    inner_raw = inner_raw[:MAX_DOWNLOAD_BYTES]

    result = _parse_csv_bytes(inner_raw, max_rows, truncated=truncated)
    result["format"] = "zip"
    result["member_name"] = member_name
    return result


def format_table(headers: list[str], rows: list[list[str]], max_col_width: int = 40) -> str:
    """Format headers and rows as a readable text table."""
    if not headers:
        return "No data available."

    def trunc(val: str) -> str:
        val = val.strip()
        return val[:max_col_width] + "..." if len(val) > max_col_width else val

    display_headers = [trunc(h) for h in headers]
    display_rows = [[trunc(cell) for cell in row] for row in rows]

    ncols = len(display_headers)
    for row in display_rows:
        while len(row) < ncols:
            row.append("")

    col_widths = [len(h) for h in display_headers]
    for row in display_rows:
        for i, cell in enumerate(row[:ncols]):
            col_widths[i] = max(col_widths[i], len(cell))

    def fmt_row(cells: list[str]) -> str:
        return " | ".join(
            cell.ljust(col_widths[i]) for i, cell in enumerate(cells[:ncols])
        )

    lines = [fmt_row(display_headers)]
    lines.append("-+-".join("-" * w for w in col_widths))
    for row in display_rows:
        lines.append(fmt_row(row))

    return "\n".join(lines)
