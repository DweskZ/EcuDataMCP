"""Client for the BCE's *Información Estadística Mensual* (IEM).

The BCEData API is excellent for a compact set of ready-made time series.
IEM is different: its monthly bulletin pages link to dozens of individual
XLSX tables, including detailed cuts that do not appear in BCEData.  This
module indexes those links live; it deliberately does not guess file paths
from a bulletin number because the BCE occasionally points a new bulletin at
the previous month's directory.
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import json
import logging
import os
import re
import zipfile
from datetime import UTC, datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import httpx
from openpyxl import load_workbook

from helpers.cache import TtlCache
from helpers.csv_reader import download_bytes
from helpers.logging import MAIN_LOGGER_NAME
from helpers.safe_download import safe_stream
from helpers.text_utils import strip_accents as _strip
from helpers.user_agent import USER_AGENT

logger = logging.getLogger(MAIN_LOGGER_NAME)

IEM_INDEX_URL = (
    "https://contenido.bce.fin.ec/documentos/informacioneconomica/"
    "PublicacionesGenerales/IndiceIEM.html"
)
IEM_LATEST_PUBLICATIONS_URL = "https://contenido.bce.fin.ec/ultimas-publicaciones/"
IEM_ARCHIVE_URL = "https://contenido.bce.fin.ec/iem-publicaciones/"
_SOURCE_NAME = "Banco Central del Ecuador — Información Estadística Mensual"
_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_CATALOG_DIR = _ROOT / "data" / "iem_catalog_snapshots"

_catalog_cache = TtlCache(ttl_seconds=86400.0, max_entries=1)
_bulletins_cache = TtlCache(ttl_seconds=86400.0, max_entries=1)
_bulletin_tables_cache = TtlCache(ttl_seconds=86400.0, max_entries=512)
# Legacy bulletin ZIPs (pre-Oct-2016) are fetched whole -- both to list their
# members (building the table catalog) and later to read one member's bytes.
# Caching avoids a second ~1-2 MB download for the very common "list, then
# read one of the tables just listed" sequence.
_zip_bytes_cache = TtlCache(ttl_seconds=86400.0, max_entries=16)
_fetch_lock = asyncio.Lock()
# Keyed per bulletin number so concurrent historical fetches (bounded by
# _HISTORY_CONCURRENCY below) actually run in parallel instead of serializing
# on one shared lock. Safe to populate via plain dict access -- asyncio is
# single-threaded and no await happens between the get and the set.
_bulletin_fetch_locks: dict[int, asyncio.Lock] = {}
_HISTORY_CONCURRENCY = 12
# Caps the fan-out for an unbounded historico=True search (no desde_anio/
# hasta_anio): otherwise every bulletin the IEM index has ever listed would
# be fetched in one call. ~5 years of monthly bulletins.
_MAX_HISTORICAL_BULLETINS = 60
_HASH_CONCURRENCY = 4
_MAX_HASH_FILES = 5000


def _bulletin_lock(key: int) -> asyncio.Lock:
    lock = _bulletin_fetch_locks.get(key)
    if lock is None:
        lock = asyncio.Lock()
        _bulletin_fetch_locks[key] = lock
    return lock
_LINK_RE = re.compile(
    r'<a\s+[^>]*href=["\'](?P<url>[^"\']+)["\'][^>]*>(?P<label>.*?)</a>',
    re.IGNORECASE | re.DOTALL,
)
_BULLETIN_RE = re.compile(
    r"m(?P<number>\d{4})(?P<month>\d{2})(?P<year>\d{4})\.html?(?:$|[?#])",
    re.IGNORECASE,
)
_TABLE_RE = re.compile(
    r"(?P<name>IEM-(?P<number>\d+[\w-]*)-e)\.xlsx$", re.IGNORECASE
)
_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")

# Bulletins before No. 1976 (October 2016) never link individual XLSX on the
# bulletin page at all -- confirmed live 2026-09-02 there are actually two
# older eras, not one:
#   - No. 1854-1975 (Aug 2006 - Sep 2016, ~122 bulletins): modern-ish
#     href="..." markup, bulk-ZIP-only. The ZIP already has one legacy .xls
#     per table, same IEM-{n} numbering as today, just packaged differently
#     -- not an unknown format. Handled by _fetch_legacy_zip_tables below.
#   - No. 1727-1853 (Jan 1996 - Jul 2006, ~126 bulletins): pre-modern-HTML
#     frameset pages (uppercase, unquoted <A HREF = ... TARGET="_top">,
#     confirmed live on No. 1800/1780) that link to ~60 per-SECTION .htm
#     pages instead of a table or a ZIP directly. The table data lives
#     nowhere else -- it's a raw HTML <TABLE> embedded in each section
#     page, with real ROWSPAN/COLSPAN multi-level headers and no closing
#     </TR>/</TH>/</TD> tags (implicit-close markup, pre-XHTML). Handled
#     by _fetch_legacy_frameset_tables/_TableGridParser below as a
#     layout-preserving grid (like the existing "vista" fallback for
#     unrecognized XLSX shapes) -- the header hierarchy is genuinely
#     irregular enough across bulletins that guessing a semantic
#     wide/long shape isn't attempted, same philosophy as _inspect_xlsx.
_LEGACY_XLS_MEMBER_RE = re.compile(r"\.xls$", re.IGNORECASE)
_LEGACY_FRAMESET_LINK_RE = re.compile(r"<A\s+HREF\s*=\s*(?P<url>[^\s>]+)", re.IGNORECASE)
_LEGACY_FRAMESET_SECTION_RE = re.compile(r"/m\d+_\d+\.htm$", re.IGNORECASE)
_LEGACY_FRAMESET_TITLE_RE = re.compile(r"<H\d[^>]*>(?P<title>.*?)</H\d>", re.IGNORECASE | re.DOTALL)


def _clean(text: str) -> str:
    return _WS_RE.sub(" ", unescape(_TAG_RE.sub(" ", text))).strip()


def _parse_bulletins(html: str) -> list[dict[str, Any]]:
    """Extract monthly bulletin pages from the IEM index, newest first."""
    bulletins: list[dict[str, Any]] = []
    for match in _LINK_RE.finditer(html):
        raw_url = unescape(match.group("url"))
        parsed = _BULLETIN_RE.search(raw_url.split("?", 1)[0])
        if parsed is None:
            continue
        bulletins.append(
            {
                "numero": int(parsed.group("number")),
                "mes": int(parsed.group("month")),
                "anio": int(parsed.group("year")),
                "titulo": _clean(match.group("label")),
                "url": urljoin(IEM_INDEX_URL, raw_url),
            }
        )
    return sorted(bulletins, key=lambda item: item["numero"], reverse=True)


def _merge_bulletins(*bulletin_lists: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Merge bulletin discoveries, preferring the first occurrence per number."""
    merged: dict[int, dict[str, Any]] = {}
    for bulletins in bulletin_lists:
        for bulletin in bulletins:
            merged.setdefault(bulletin["numero"], bulletin)
    return sorted(merged.values(), key=lambda item: item["numero"], reverse=True)


def _bulletin_diagnostics(bulletins: list[dict[str, Any]]) -> dict[str, Any]:
    """Summarize the discovered archive without pretending numbers are dates."""
    numbers = sorted({int(item["numero"]) for item in bulletins})
    gaps = [number for number in range(numbers[0], numbers[-1] + 1) if number not in numbers]
    return {
        "boletines_descubiertos": len(numbers),
        "primer_boletin": numbers[0] if numbers else None,
        "ultimo_boletin": numbers[-1] if numbers else None,
        "numeros_faltantes": gaps,
    }


def _parse_complete_files(html: str, bulletin: dict[str, Any]) -> list[dict[str, Any]]:
    """Catalog complete PDF/ZIP downloads linked by one bulletin page."""
    files: list[dict[str, Any]] = []
    seen: set[str] = set()
    for match in _LINK_RE.finditer(html):
        raw_url = unescape(match.group("url"))
        filename = raw_url.split("?", 1)[0].rsplit("/", 1)[-1]
        suffix = filename.lower()
        if not suffix.endswith((".pdf", ".zip")):
            continue
        url = urljoin(bulletin["url"], raw_url)
        if url in seen:
            continue
        seen.add(url)
        files.append(
            {
                "nombre": filename,
                "tipo": "pdf" if suffix.endswith(".pdf") else "zip",
                "titulo": _clean(match.group("label")) or filename,
                "url": url,
                "boletin_numero": bulletin["numero"],
                "boletin_mes": bulletin["mes"],
                "boletin_anio": bulletin["anio"],
            }
        )
    return files


def _section_before(html: str, position: int) -> str:
    """Best-effort section label immediately before a table link.

    IEM is hand-authored HTML rather than an API.  The table title is the
    authoritative searchable label; this breadcrumb is only helpful context.
    """
    nearby = _clean(html[max(0, position - 900) : position])
    matches = re.findall(r"(?:^|\s)([1-4](?:\.\d+){0,2}\s+[^\n]+)", nearby)
    return matches[-1].strip() if matches else ""


def _parse_tables(html: str, bulletin: dict[str, Any]) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    seen: set[str] = set()
    for match in _LINK_RE.finditer(html):
        raw_url = unescape(match.group("url"))
        filename = raw_url.split("?", 1)[0].rsplit("/", 1)[-1]
        file_match = _TABLE_RE.search(filename)
        if file_match is None:
            continue
        table_id = file_match.group("name").lower()
        if table_id in seen:
            continue
        seen.add(table_id)
        tables.append(
            {
                "table_id": table_id,
                "titulo": _clean(match.group("label")) or table_id,
                "seccion": _section_before(html, match.start()),
                "url": urljoin(bulletin["url"], raw_url),
                "boletin_numero": bulletin["numero"],
                "boletin_mes": bulletin["mes"],
                "boletin_anio": bulletin["anio"],
                "boletin_url": bulletin["url"],
                "catalogado_en": datetime.now(UTC).isoformat(),
            }
        )
    return tables


async def _fetch_catalog() -> dict[str, Any]:
    """Fetch the current bulletin catalogue (the cheap/default path)."""
    cached = _catalog_cache.get("catalog")
    if cached is not None:
        return cached
    bulletins = await _fetch_bulletins()
    bulletin = bulletins[0]
    tables = await _fetch_tables_for_bulletin(bulletin)
    return _catalog_result(bulletin, tables, **_bulletin_diagnostics(bulletins))


async def _fetch_bulletins() -> list[dict[str, Any]]:
    cached = _bulletins_cache.get("bulletins")
    if cached is not None:
        return cached

    async with _fetch_lock:
        cached = _bulletins_cache.get("bulletins")
        if cached is not None:
            return cached

        raw_index, truncated = await download_bytes(IEM_INDEX_URL)
        if truncated:
            raise ValueError("El índice IEM del BCE superó el límite de descarga")
        index_bulletins = _parse_bulletins(
            raw_index.decode("utf-8", errors="replace")
        )
        latest_bulletins: list[dict[str, Any]] = []
        archive_bulletins: list[dict[str, Any]] = []
        try:
            raw_latest, latest_truncated = await download_bytes(
                IEM_LATEST_PUBLICATIONS_URL
            )
            if not latest_truncated:
                latest_bulletins = _parse_bulletins(
                    raw_latest.decode("utf-8", errors="replace")
                )
        except Exception as exc:
            # Keep the historical index as a safe fallback if the publications
            # page is temporarily unavailable or changes its markup.
            logger.warning(
                "No se pudo reconciliar Últimas publicaciones del BCE: %s", exc
            )
        try:
            raw_archive, archive_truncated = await download_bytes(IEM_ARCHIVE_URL)
            if not archive_truncated:
                archive_bulletins = _parse_bulletins(
                    raw_archive.decode("utf-8", errors="replace")
                )
        except Exception as exc:
            # The older archive is supplementary: preserve the current index
            # when it is temporarily unavailable or its WordPress markup moves.
            logger.warning("No se pudo cargar el archivo histórico IEM del BCE: %s", exc)
        bulletins = _merge_bulletins(
            latest_bulletins, index_bulletins, archive_bulletins
        )
        if not bulletins:
            raise ValueError("No se encontraron boletines IEM en el índice del BCE")
        _bulletins_cache.set("bulletins", bulletins)
        return bulletins


def _legacy_table_id(filename: str) -> str:
    """table_id for one member of a pre-Oct-2016 bulk ZIP.

    Deliberately a different shape from _TABLE_RE's "iem-NNN-e": whether a
    legacy "IEM-315.xls" is the same table as the modern "iem-315-e" is not
    confirmed (BCE's own section numbering could have shifted across a
    decade), and some legacy members don't even follow the IEM-NNN pattern
    at all (e.g. "5_SectorPetrolero.xls", "7_GraficosIDEAC.xls"). Prefixing
    every legacy id avoids silently implying an equivalence this client
    hasn't verified.
    """
    stem = re.sub(r"\.xls$", "", filename, flags=re.IGNORECASE)
    normalized = re.sub(r"[^a-z0-9]+", "-", stem.lower()).strip("-")
    return f"iem-legado-{normalized}"


async def _download_zip_cached(url: str) -> bytes:
    cached = _zip_bytes_cache.get(url)
    if cached is not None:
        return cached
    raw, truncated = await download_bytes(url)
    if truncated:
        raise ValueError("El ZIP del boletín IEM superó el límite de descarga")
    _zip_bytes_cache.set(url, raw)
    return raw


async def _fetch_legacy_zip_tables(
    bulletin: dict[str, Any], zip_info: dict[str, Any]
) -> list[dict[str, Any]]:
    """Build a table catalog for a bulletin whose page has no individual
    XLSX links, from the bulk ZIP's own member list instead."""
    raw = await _download_zip_cached(zip_info["url"])
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile as exc:
        raise ValueError(
            f"El ZIP del boletín IEM {bulletin['numero']} no es un archivo válido"
        ) from exc
    tables: list[dict[str, Any]] = []
    for name in archive.namelist():
        if not _LEGACY_XLS_MEMBER_RE.search(name):
            continue
        tables.append(
            {
                "table_id": _legacy_table_id(name),
                "titulo": name,
                "seccion": "",
                "url": zip_info["url"],
                "zip_member": name,
                "formato_origen": "xls_legado_zip",
                "boletin_numero": bulletin["numero"],
                "boletin_mes": bulletin["mes"],
                "boletin_anio": bulletin["anio"],
                "boletin_url": bulletin["url"],
                "catalogado_en": datetime.now(UTC).isoformat(),
            }
        )
    return tables


class _TableGridParser(HTMLParser):
    """Collect one or more HTML <TABLE>s worth of (text, rowspan, colspan)
    cells, tolerating the pre-XHTML BCE markup this era actually uses: no
    closing </TR>/</TH>/</TD> at all. A tokenizer, not a tree builder, so
    it replicates the browser's implicit-close inference itself -- a new
    TR/TH/TD start tag ends whatever same-level element was open before
    it, exactly like handle_endtag would for well-formed markup.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[tuple[str, int, int]]]] = []
        self._in_table = False
        self._rows: list[list[tuple[str, int, int]]] | None = None
        self._row: list[tuple[str, int, int]] | None = None
        self._cell: list[str] | None = None
        self._cell_span = (1, 1)

    def _close_cell(self) -> None:
        if self._cell is None or self._row is None:
            return
        text = _WS_RE.sub(" ", "".join(self._cell)).strip()
        rowspan, colspan = self._cell_span
        self._row.append((text, max(rowspan, 1), max(colspan, 1)))
        self._cell = None

    def _close_row(self) -> None:
        self._close_cell()
        if self._row is not None and self._rows is not None:
            self._rows.append(self._row)
        self._row = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._in_table = True
            self._rows = []
            self._row = None
            self._cell = None
            return
        if not self._in_table:
            return
        if tag == "tr":
            self._close_row()
            self._row = []
        elif tag in {"th", "td"}:
            self._close_cell()
            if self._row is None:
                self._row = []
            values = dict(attrs)
            try:
                rowspan = int(values.get("rowspan") or 1)
            except ValueError:
                rowspan = 1
            try:
                colspan = int(values.get("colspan") or 1)
            except ValueError:
                colspan = 1
            self._cell_span = (rowspan, colspan)
            self._cell = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "table" and self._in_table:
            self._close_row()
            if self._rows:
                self.tables.append(self._rows)
            self._in_table = False
            self._rows = None

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def _expand_table_grid(rows: list[list[tuple[str, int, int]]]) -> list[list[str]]:
    """Resolve ROWSPAN/COLSPAN cells into a plain rectangular grid.

    Standard algorithm: track cells still "hanging" into future rows from
    an earlier ROWSPAN, filling them in before placing each row's own new
    cells at the first free column.
    """
    grid: list[list[str]] = []
    carry: dict[int, tuple[int, str]] = {}  # column -> (rows_remaining, text)
    for row in rows:
        col = 0
        out_row: list[str] = []

        def fill_carried(col: int, out_row: list[str]) -> int:
            while col in carry:
                remaining, text = carry[col]
                out_row.append(text)
                if remaining > 1:
                    carry[col] = (remaining - 1, text)
                else:
                    del carry[col]
                col += 1
            return col

        col = fill_carried(col, out_row)
        for text, rowspan, colspan in row:
            col = fill_carried(col, out_row)
            for offset in range(colspan):
                out_row.append(text)
                if rowspan > 1:
                    carry[col + offset] = (rowspan - 1, text)
            col += colspan
        grid.append(out_row)
    return grid


def _legacy_frameset_table_id(title: str, fallback_index: int) -> str:
    """table_id for a frameset-era section table, derived from its own
    title text rather than the m{boletin}_{k}.htm index -- that index
    isn't confirmed stable across bulletins, but the section title (e.g.
    "1.1 Principales Indicadores Monetarios") is the same kind of stable,
    human-assigned label BCE still uses today."""
    normalized = re.sub(r"[^a-z0-9]+", "-", _strip(title).lower()).strip("-")
    if not normalized:
        normalized = f"seccion-{fallback_index}"
    return f"iem-legado-frameset-{normalized}"


async def _fetch_legacy_frameset_tables(
    bulletin: dict[str, Any], html: str
) -> list[dict[str, Any]]:
    """Discover section pages linked from a pre-Aug-2006 bulletin's
    uppercase/unquoted <A HREF = ...> frameset markup, which
    _parse_tables/_parse_complete_files' href="..." regex never matches."""
    section_urls: list[str] = []
    seen: set[str] = set()
    for match in _LEGACY_FRAMESET_LINK_RE.finditer(html):
        raw_url = match.group("url").strip("'\"")
        if not _LEGACY_FRAMESET_SECTION_RE.search(raw_url):
            continue
        url = urljoin(bulletin["url"], raw_url)
        if url in seen:
            continue
        seen.add(url)
        section_urls.append(url)

    tables: list[dict[str, Any]] = []
    for index, url in enumerate(section_urls, start=1):
        try:
            raw_section, truncated = await download_bytes(url)
        except Exception as exc:
            logger.warning("No se pudo leer la sección %s del boletín IEM: %s", url, exc)
            continue
        if truncated:
            continue
        section_html = raw_section.decode("cp1252", errors="replace")
        title_match = _LEGACY_FRAMESET_TITLE_RE.search(section_html)
        title = _clean(unescape(title_match.group("title"))) if title_match else f"Sección {index}"
        tables.append(
            {
                "table_id": _legacy_frameset_table_id(title, index),
                "titulo": title or f"Sección {index}",
                "seccion": "",
                "url": url,
                "formato_origen": "html_frameset",
                "boletin_numero": bulletin["numero"],
                "boletin_mes": bulletin["mes"],
                "boletin_anio": bulletin["anio"],
                "boletin_url": bulletin["url"],
                "catalogado_en": datetime.now(UTC).isoformat(),
            }
        )
    return tables


async def _fetch_tables_for_bulletin(bulletin: dict[str, Any]) -> list[dict[str, Any]]:
    key = bulletin["numero"]
    cached = _bulletin_tables_cache.get(key)
    if cached is not None:
        return cached

    async with _bulletin_lock(key):
        cached = _bulletin_tables_cache.get(key)
        if cached is not None:
            return cached

        raw_bulletin, truncated = await download_bytes(bulletin["url"])
        if truncated:
            raise ValueError("La página del boletín IEM superó el límite de descarga")
        html = raw_bulletin.decode("utf-8", errors="replace")
        bulletin["archivos_completos"] = _parse_complete_files(html, bulletin)
        tables = _parse_tables(html, bulletin)
        if not tables:
            zip_info = next(
                (f for f in bulletin["archivos_completos"] if f["tipo"] == "zip"), None
            )
            if zip_info is not None:
                tables = await _fetch_legacy_zip_tables(bulletin, zip_info)
        if not tables:
            tables = await _fetch_legacy_frameset_tables(bulletin, html)
        if not tables:
            raise ValueError(
                f"El boletín IEM {bulletin['numero']} no expuso tablas XLSX individuales"
            )
        _bulletin_tables_cache.set(key, tables)
        return tables


def _build_catalog(
    bulletin: dict[str, Any], tables: list[dict[str, Any]], **extra: Any
) -> dict[str, Any]:
    return {
        "source": _SOURCE_NAME,
        "url_fuente": IEM_INDEX_URL,
        "boletin": bulletin,
        "catalogado_en": datetime.now(UTC).isoformat(),
        **extra,
        "total_tablas": len(tables),
        "tablas": tables,
    }


def _catalog_result(
    bulletin: dict[str, Any], tables: list[dict[str, Any]], **extra: Any
) -> dict[str, Any]:
    result = _build_catalog(bulletin, tables, **extra)
    _catalog_cache.set("catalog", result)
    logger.info("IEM BCE indexado: boletín %d, %d tablas", bulletin["numero"], len(tables))
    return result


def _iem_catalog_dir() -> Path:
    configured = os.getenv("IEM_CATALOG_DIR", "").strip()
    return Path(configured) if configured else _DEFAULT_CATALOG_DIR


def persist_catalog(catalog: dict[str, Any]) -> dict[str, Any]:
    """Persist a complete IEM catalog atomically for later inspection."""
    directory = _iem_catalog_dir()
    directory.mkdir(parents=True, exist_ok=True)
    now = datetime.now(UTC)
    stamp = re.sub(r"[^0-9A-Za-z_.-]+", "-", now.isoformat()).strip("-")
    payload = {
        **catalog,
        "persistido_en": now.isoformat(),
        "catalogo_id": stamp,
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2).encode(
        "utf-8"
    )
    path = directory / f"catalog-{stamp}.json"
    temporary = directory / f".{path.name}.tmp"
    temporary.write_bytes(raw)
    temporary.replace(path)
    latest = directory / "latest.json"
    temporary_latest = directory / ".latest.json.tmp"
    temporary_latest.write_bytes(raw)
    temporary_latest.replace(latest)
    return {"directorio": str(directory), "archivo": str(path)}


def _within_years(
    bulletin: dict[str, Any], desde_anio: int, hasta_anio: int
) -> bool:
    return (
        (not desde_anio or bulletin["anio"] >= desde_anio)
        and (not hasta_anio or bulletin["anio"] <= hasta_anio)
    )


async def _fetch_historical_tables(
    bulletins: list[dict[str, Any]], desde_anio: int, hasta_anio: int
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    """Fetch tables for every bulletin in range, capped only when unbounded.

    Returns (tables, selected_bulletins, loaded_count). A caller who passes
    an explicit desde_anio/hasta_anio is trusted to have bounded their own
    request; only a bare historico=True (no bounds at all) is capped, since
    that would otherwise fetch every bulletin ever published. ``bulletins``
    is already newest-first, so capping keeps the most recent ones.
    """
    selected_bulletins = [
        bulletin
        for bulletin in bulletins
        if _within_years(bulletin, desde_anio, hasta_anio)
    ]
    if not selected_bulletins:
        raise ValueError("No hay boletines IEM en el rango solicitado")
    if not desde_anio and not hasta_anio:
        selected_bulletins = selected_bulletins[:_MAX_HISTORICAL_BULLETINS]

    semaphore = asyncio.Semaphore(_HISTORY_CONCURRENCY)

    async def fetch(bulletin: dict[str, Any]) -> list[dict[str, Any]]:
        async with semaphore:
            return await _fetch_tables_for_bulletin(bulletin)

    batches = await asyncio.gather(
        *(fetch(b) for b in selected_bulletins), return_exceptions=True
    )
    tables: list[dict[str, Any]] = []
    loaded = 0
    for bulletin, batch in zip(selected_bulletins, batches, strict=True):
        if isinstance(batch, BaseException):
            logger.warning(
                "No se pudo indexar el boletín IEM %d: %s",
                bulletin["numero"],
                batch,
            )
            continue
        loaded += 1
        tables.extend(batch)
    if not tables:
        raise ValueError("Ningún boletín IEM del rango expuso tablas XLSX")
    return tables, selected_bulletins, loaded


def _merge_historical_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Present one result per table, retaining every bulletin version."""
    grouped: dict[str, list[dict[str, Any]]] = {}
    for table in tables:
        grouped.setdefault(table["table_id"], []).append(table)

    merged = []
    for versions in grouped.values():
        versions.sort(key=lambda item: item["boletin_numero"], reverse=True)
        current = dict(versions[0])
        current["versiones"] = [
            {
                "boletin_numero": version["boletin_numero"],
                "boletin_mes": version["boletin_mes"],
                "boletin_anio": version["boletin_anio"],
                "url": version["url"],
            }
            for version in versions
        ]
        current["boletines_disponibles"] = len(versions)
        merged.append(current)
    return sorted(merged, key=lambda item: item["titulo"].lower())


async def search_tables(
    query: str = "",
    limit: int = 20,
    offset: int = 0,
    historico: bool = False,
    desde_anio: int = 0,
    hasta_anio: int = 0,
    guardar_catalogo: bool = False,
    hash_archivos: bool = False,
    max_hash_archivos: int = _MAX_HASH_FILES,
) -> dict[str, Any]:
    """Search current tables, or versions across the IEM bulletin archive."""
    if desde_anio and hasta_anio and desde_anio > hasta_anio:
        raise ValueError("desde_anio no puede ser mayor que hasta_anio")
    hash_targets: list[dict[str, Any]] | None = None
    if historico or desde_anio or hasta_anio:
        bulletins = await _fetch_bulletins()
        all_tables, selected_bulletins, loaded_bulletins = await _fetch_historical_tables(
            bulletins, desde_anio, hasta_anio
        )
        hash_targets = all_tables
        tables = _merge_historical_tables(all_tables)
        catalog = _build_catalog(
            selected_bulletins[0],
            tables,
            boletines_consultados=loaded_bulletins,
            boletines_sin_tablas=len(selected_bulletins) - loaded_bulletins,
            **_bulletin_diagnostics(bulletins),
        )
    else:
        catalog = await _fetch_catalog()
        hash_targets = catalog["tablas"]
    hash_result = None
    if hash_archivos:
        hash_result = await hash_catalog_tables(hash_targets or [], max_hash_archivos)
    if guardar_catalogo:
        catalogo = persist_catalog(catalog)
    else:
        catalogo = None
    q = _strip(query)
    tables = catalog["tablas"]
    matched = [
        table
        for table in tables
        if not q or q in _strip(f"{table['titulo']} {table['seccion']} {table['table_id']}")
    ]
    result = {
        **{key: value for key, value in catalog.items() if key != "tablas"},
        "total": len(matched),
        "offset": offset,
        "historico": historico or bool(desde_anio or hasta_anio),
        "tablas": matched[offset : offset + limit],
    }
    if catalogo is not None:
        result["catalogo_guardado"] = catalogo
    if hash_result is not None:
        result["hash_archivos"] = hash_result
    return result


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _trim_trailing_blanks(row: list[str]) -> list[str]:
    while row and not row[-1]:
        row.pop()
    return row


def _year(value: Any) -> int | None:
    match = re.match(r"\s*(\d{4})", str(value or ""))
    return int(match.group(1)) if match else None


_MONTHS = {
    "ene": 1, "enero": 1, "feb": 2, "febrero": 2,
    "mar": 3, "marzo": 3, "abr": 4, "abril": 4,
    "may": 5, "mayo": 5, "jun": 6, "junio": 6,
    "jul": 7, "julio": 7, "ago": 8, "agosto": 8,
    "sep": 9, "sept": 9, "septiembre": 9, "oct": 10,
    "octubre": 10, "nov": 11, "noviembre": 11, "dic": 12,
    "diciembre": 12,
}


def _period_key(value: Any) -> tuple[int, int] | None:
    """Normalize annual, monthly and Spanish month-year labels for filtering."""
    if isinstance(value, datetime):
        return value.year, value.month
    text = _strip(str(value or "")).strip().replace(".", "")
    year_match = re.search(r"(?<!\d)(\d{4})(?!\d)", text)
    if year_match is None:
        return None
    year = int(year_match.group(1))
    year_month = re.fullmatch(r"\d{4}[-/](0?[1-9]|1[0-2])", text)
    if year_month:
        return year, int(year_month.group(1))
    month_year = re.fullmatch(r"(0?[1-9]|1[0-2])[-/]\d{4}", text)
    if month_year:
        return year, int(month_year.group(1))
    compact_quarter = re.fullmatch(r"\d{4}\s*(?:t|q)\s*([1-4])", text, re.IGNORECASE)
    if compact_quarter:
        return year, (int(compact_quarter.group(1)) - 1) * 3 + 1
    prefix = text[: year_match.start()].strip(" -/_")
    suffix = text[year_match.end() :].strip(" -/_")
    month_name = prefix or suffix
    if month_name in _MONTHS:
        return year, _MONTHS[month_name]
    quarter = re.search(
        r"(?<!\d)([1-4])\s*(?:er|ro|do|to)?\s*(?:trimestre|trim\.?|t)(?:\s|[-/])*"
        r"(\d{4})(?!\d)",
        text,
        re.IGNORECASE,
    )
    if quarter:
        return int(quarter.group(2)), (int(quarter.group(1)) - 1) * 3 + 1
    roman_quarter = re.search(
        r"\b((?:i{1,3}|iv))\s*(?:trimestre|trim\.?)"
        r"(?:\s|[-/])*(\d{4})\b",
        text,
        re.IGNORECASE,
    )
    if roman_quarter:
        quarter_number = {"i": 1, "ii": 2, "iii": 3, "iv": 4}[roman_quarter.group(1).lower()]
        return int(roman_quarter.group(2)), (quarter_number - 1) * 3 + 1
    numeric_month = re.search(r"(?<!\d)(0?[1-9]|1[0-2])(?=[/-])", text)
    return (year, int(numeric_month.group(1))) if numeric_month else (year, 0)


def _period_bounds(
    desde: str, hasta: str
) -> tuple[tuple[int, int] | None, tuple[int, int] | None]:
    start = _period_key(desde) if desde else None
    end = _period_key(hasta) if hasta else None
    if desde and start is None:
        raise ValueError("desde debe ser un período YYYY, YYYY-MM o mes-año")
    if hasta and end is None:
        raise ValueError("hasta debe ser un período YYYY, YYYY-MM o mes-año")
    if start and end:
        start_cmp = (start[0], start[1] or 1)
        end_cmp = (end[0], end[1] or 12)
        if start_cmp > end_cmp:
            raise ValueError("desde no puede ser mayor que hasta")
    return start, end


def _period_in_bounds(
    period: tuple[int, int],
    start: tuple[int, int] | None,
    end: tuple[int, int] | None,
) -> bool:
    year, month = period
    if month == 0:
        return (start is None or year >= start[0]) and (end is None or year <= end[0])
    return (
        (start is None or (year, month) >= (start[0], start[1] or 1))
        and (end is None or (year, month) <= (end[0], end[1] or 12))
    )


def _selected_columns(
    header: tuple[Any, ...], desde: str, hasta: str
) -> list[tuple[int, str]]:
    start, end = _period_bounds(desde, hasta)

    selected = []
    for index, value in enumerate(header[1:], start=1):
        period = _period_key(value)
        if period is None or not _period_in_bounds(period, start, end):
            continue
        selected.append((index, str(value)))
    return selected


def _extract_wide_series(
    worksheet: Any, desde: str, hasta: str, max_rows: int
) -> dict[str, Any] | None:
    """Read the common IEM shape: periods across columns, series down rows.

    This is verified against IEM-431-e. Other IEM tables can differ, so a
    non-match returns None and the caller gives a layout-preserving preview.
    """
    all_rows = list(worksheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(all_rows)
            if row
            and any(
                token in _strip(str(row[0] or ""))
                for token in ("period", "ano", "fecha", "mes")
            )
            and len(_selected_columns(row, "", "")) >= 2
        ),
        None,
    )
    if header_index is None:
        return None

    selected = _selected_columns(all_rows[header_index], desde, hasta)
    if not selected:
        # The table has periods, just not any within desde/hasta -- let the
        # caller fall back to _extract_long_table / the raw preview instead
        # of failing outright.
        return None

    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in all_rows[header_index + 1 :]:
        label = str(row[0]).strip() if row and row[0] is not None else ""
        if not label:
            continue
        # Check the whole row, not just the desde/hasta-selected columns --
        # a series with no data in the requested range would otherwise look
        # like a unit-block header and swallow every row after it.
        row_values = row[1:] if row else ()
        if not any(value is not None for value in row_values):
            # A one-cell row names a unit block, such as "Millones de USD".
            current = {"unidad": label, "series": [], "truncada": False}
            blocks.append(current)
            continue
        if current is None:
            # Some monthly tables omit the unit block and start with a
            # Variable/Indicador row. Create an explicit unitless block so
            # the following series are still normalized.
            if _strip(label) in {"variable", "indicador", "serie"}:
                current = {"unidad": "", "series": [], "truncada": False}
                blocks.append(current)
                continue
            current = {"unidad": "", "series": [], "truncada": False}
            blocks.append(current)
        if len(current["series"]) >= max_rows:
            current["truncada"] = True
            continue
        current["series"].append(
            {
                "nombre": label,
                "valores": {
                    period: row[index] if index < len(row) else None
                    for index, period in selected
                },
            }
        )

    blocks = [block for block in blocks if block["series"]]
    if not blocks:
        return None
    return {
        "formato": "series_ancho",
        "periodos": [period for _, period in selected],
        "bloques": blocks,
    }


def _extract_long_table(
    worksheet: Any, desde: str, hasta: str, max_rows: int
) -> dict[str, Any] | None:
    """Read a second common shape: one observation per row.

    This handles tables whose header has columns such as Año/Período,
    Variable and Valor. It keeps the original column names and cell values;
    it does not rename fields because IEM labels are part of the source.
    """
    all_rows = list(worksheet.iter_rows(values_only=True))
    header_index = None
    header: tuple[Any, ...] | None = None
    period_index = None
    for index, row in enumerate(all_rows[:40]):
        names = [_strip(str(value or "")) for value in row]
        candidate_period = next(
            (
                position
                for position, name in enumerate(names)
                if any(
                    token in name for token in ("period", "ano", "fecha", "mes")
                )
            ),
            None,
        )
        nonempty = sum(bool(name) for name in names)
        if candidate_period is not None and nonempty >= 2:
            header_index = index
            header = row
            period_index = candidate_period
            break
    if header_index is None or header is None or period_index is None:
        return None

    start, end = _period_bounds(desde, hasta)

    headers = _trim_trailing_blanks([_cell(value) for value in header])
    rows: list[list[str]] = []
    total_rows = 0
    for raw_row in all_rows[header_index + 1 :]:
        row = [_cell(value) for value in raw_row[: len(headers)]]
        if not any(row):
            continue
        period = _period_key(
            raw_row[period_index] if period_index < len(raw_row) else None
        )
        if period is None or not _period_in_bounds(period, start, end):
            continue
        total_rows += 1
        if len(rows) < max_rows:
            rows.append(_trim_trailing_blanks(row))
    if not rows:
        return None
    return {
        "formato": "tabla_larga",
        "encabezados": headers,
        "filas": rows,
        "filas_totales": total_rows,
        "truncada": total_rows > len(rows),
    }


def _extract_matrix_series(
    worksheet: Any, desde: str, hasta: str, max_rows: int
) -> dict[str, Any] | None:
    """Normalize tables with several descriptor columns before the periods.

    Some IEM families use a matrix such as ``Código | Descripción | 2025 | …``
    or put quarterly/monthly labels in a header row that is not named
    ``Período``.  The existing wide parser intentionally stays conservative;
    this parser handles that family while preserving the source descriptors.
    """
    all_rows = list(worksheet.iter_rows(values_only=True))
    header_index = None
    period_columns: list[tuple[int, str]] = []
    for index, row in enumerate(all_rows[:50]):
        periods = [
            (position, str(value))
            for position, value in enumerate(row)
            if _period_key(value) is not None
        ]
        if len(periods) >= 2:
            header_index = index
            period_columns = periods
            break
    if header_index is None:
        return None

    start, end = _period_bounds(desde, hasta)
    selected = [
        (position, label)
        for position, label in period_columns
        if _period_in_bounds(_period_key(label) or (0, 0), start, end)
    ]
    if not selected:
        return None

    first_period_column = min(position for position, _ in period_columns)
    descriptor_columns = list(range(first_period_column))
    if not descriptor_columns:
        return None
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in all_rows[header_index + 1 :]:
        descriptors = [
            _cell(row[position]).strip()
            for position in descriptor_columns
            if position < len(row) and _cell(row[position]).strip()
        ]
        if not descriptors:
            continue
        label = " | ".join(descriptors)
        values = {
            period: row[position] if position < len(row) else None
            for position, period in selected
        }
        if not any(value is not None for value in values.values()):
            current = {"unidad": label, "series": [], "truncada": False}
            blocks.append(current)
            continue
        if current is None:
            current = {"unidad": "", "series": [], "truncada": False}
            blocks.append(current)
        if len(current["series"]) >= max_rows:
            current["truncada"] = True
            continue
        current["series"].append({"nombre": label, "valores": values})

    blocks = [block for block in blocks if block["series"]]
    if not blocks:
        return None
    return {
        "formato": "series_matriz",
        "periodos": [period for _, period in selected],
        "columnas_descriptivas": [
            _cell(all_rows[header_index][position])
            for position in descriptor_columns
        ],
        "bloques": blocks,
    }


async def _hash_url(url: str, session: httpx.AsyncClient) -> dict[str, Any]:
    """Hash one complete XLSX without applying the preview-size limit."""
    digest = hashlib.sha256()
    total = 0
    async with safe_stream(session, url, timeout=120.0) as response:
        response.raise_for_status()
        async for chunk in response.aiter_bytes(chunk_size=256 * 1024):
            digest.update(chunk)
            total += len(chunk)
    return {"sha256": digest.hexdigest(), "bytes": total}


async def hash_catalog_tables(
    tables: list[dict[str, Any]], max_files: int = _MAX_HASH_FILES
) -> dict[str, Any]:
    """Hash a bounded set of IEM XLSX links for reproducible archive audits."""
    max_files = min(max(max_files, 1), _MAX_HASH_FILES)
    selected = tables[:max_files]
    omitted = max(0, len(tables) - len(selected))
    semaphore = asyncio.Semaphore(_HASH_CONCURRENCY)

    async def hash_one(table: dict[str, Any], session: httpx.AsyncClient) -> dict[str, Any]:
        async with semaphore:
            try:
                result = await _hash_url(table["url"], session)
                return {**table, **result, "ok": True}
            except Exception as exc:
                return {**table, "ok": False, "error": str(exc)}

    async with httpx.AsyncClient(
        headers={"User-Agent": USER_AGENT}, timeout=120.0
    ) as session:
        results = await asyncio.gather(*(hash_one(table, session) for table in selected))
    return {
        "total_archivos": len(tables),
        "archivos_consultados": len(selected),
        "archivos_omitidos_por_limite": omitted,
        "archivos_exitosos": sum(1 for result in results if result["ok"]),
        "archivos_con_error": sum(1 for result in results if not result["ok"]),
        "resultados": results,
    }


def _inspect_worksheet_rows(rows_iterable: Any, max_rows: int) -> dict[str, Any]:
    """Shared row-collection logic behind _inspect_xlsx/_inspect_legacy_xls."""
    rows: list[list[str]] = []
    nonempty_total = 0
    for row in rows_iterable:
        rendered = _trim_trailing_blanks([_cell(value) for value in row])
        if not any(rendered):
            continue
        nonempty_total += 1
        if len(rows) < max_rows:
            rows.append(rendered[:30])
    return {
        "vista": rows,
        "filas_mostradas": len(rows),
        "truncada": nonempty_total > len(rows),
    }


def _inspect_xlsx(raw: bytes, max_rows: int) -> dict[str, Any]:
    """Return a layout-preserving preview without guessing a universal header.

    Individual IEM tables use different layouts.  Returning the first
    non-empty rows is safer than pretending a title row is always a header;
    dedicated parsers can be added later for high-value tables.
    """
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheets = [
            {
                "nombre": worksheet.title,
                **_inspect_worksheet_rows(worksheet.iter_rows(values_only=True), max_rows),
            }
            for worksheet in workbook.worksheets
        ]
        return {"hojas": sheets}
    finally:
        workbook.close()


class _XlsSheetAdapter:
    """Minimal openpyxl-worksheet-shaped wrapper around an xlrd Sheet.

    _extract_wide_series/_extract_long_table/_extract_matrix_series only
    ever call worksheet.iter_rows(values_only=True), so this one method is
    the entire surface needed to reuse them unchanged for legacy .xls
    tables pulled out of pre-Oct-2016 bulletin ZIPs -- normalizing xlrd's
    "" empty-cell value to None (openpyxl's convention, which the
    extraction functions' `is not None` checks depend on) and its raw date
    floats to datetime (which _period_key already handles).
    """

    def __init__(self, sheet: Any, book: Any) -> None:
        self._sheet = sheet
        self._book = book

    def iter_rows(self, values_only: bool = True):
        import xlrd

        for row_index in range(self._sheet.nrows):
            row: list[Any] = []
            for col_index in range(self._sheet.ncols):
                cell = self._sheet.cell(row_index, col_index)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK) or cell.value == "":
                    row.append(None)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate.xldate_as_datetime(cell.value, self._book.datemode))
                elif (
                    cell.ctype == xlrd.XL_CELL_NUMBER
                    and isinstance(cell.value, float)
                    and cell.value.is_integer()
                ):
                    # xlrd has no separate int type -- every number comes back
                    # as float. A year header like 2025.0 stringifies to
                    # "2025.0", and _period_key's str(...).replace(".", "")
                    # turns that into "20250", breaking its 4-digit year
                    # regex. openpyxl gives plain ints for whole numbers;
                    # matching that here is what makes the shared
                    # _extract_* functions and _period_key work unchanged.
                    row.append(int(cell.value))
                else:
                    row.append(cell.value)
            yield tuple(row)


def _inspect_legacy_xls(raw: bytes, max_rows: int) -> dict[str, Any]:
    """Layout-preserving preview for a legacy .xls table -- see _inspect_xlsx."""
    import xlrd

    book = xlrd.open_workbook(file_contents=raw)
    sheets = [
        {
            "nombre": sheet.name,
            **_inspect_worksheet_rows(
                _XlsSheetAdapter(sheet, book).iter_rows(), max_rows
            ),
        }
        for sheet in book.sheets()
    ]
    return {"hojas": sheets}


def _inspect_frameset_table(raw: bytes, max_rows: int) -> dict[str, Any]:
    """Layout-preserving preview for a frameset-era section page's raw HTML
    <TABLE>(s) -- see _inspect_xlsx. Never attempted as wide/long/matrix:
    this era's header hierarchies (nested ROWSPAN/COLSPAN groups, footnote
    markers mixed into header text) are genuinely irregular enough that
    guessing a semantic shape isn't safe -- the resolved grid is the
    honest result.
    """
    html = raw.decode("cp1252", errors="replace")
    parser = _TableGridParser()
    parser.feed(html)
    sheets = []
    for index, rows in enumerate(parser.tables, start=1):
        grid = [row for row in _expand_table_grid(rows) if any(cell for cell in row)]
        sheets.append(
            {
                "nombre": f"Tabla {index}",
                "vista": grid[:max_rows],
                "filas_mostradas": min(len(grid), max_rows),
                "truncada": len(grid) > max_rows,
            }
        )
    return {"hojas": sheets}


def clear_caches() -> None:
    """Clear IEM discovery caches; useful for refresh jobs and tests."""
    _catalog_cache.clear()
    _bulletins_cache.clear()
    _bulletin_tables_cache.clear()
    _bulletin_fetch_locks.clear()
    _zip_bytes_cache.clear()


async def get_table(
    table_id: str,
    desde: str = "",
    hasta: str = "",
    max_rows: int = 20,
    boletin_numero: int = 0,
) -> dict[str, Any]:
    """Download one IEM table version and return structured data when safe."""
    if boletin_numero:
        bulletins = await _fetch_bulletins()
        bulletin = next(
            (item for item in bulletins if item["numero"] == boletin_numero), None
        )
        if bulletin is None:
            raise ValueError(f"Boletín IEM '{boletin_numero}' no encontrado")
        tables = await _fetch_tables_for_bulletin(bulletin)
        catalog = _build_catalog(bulletin, tables)
    else:
        catalog = await _fetch_catalog()
    wanted = table_id.strip().lower()
    table = next((item for item in catalog["tablas"] if item["table_id"] == wanted), None)
    if table is None:
        raise ValueError(f"Tabla IEM '{table_id}' no encontrada en el boletín vigente")

    if table.get("zip_member"):
        # Pre-Oct-2016 bulletin: no individual XLSX URL, read one member
        # out of the bulk ZIP instead (see _fetch_legacy_zip_tables).
        import xlrd

        zip_raw = await _download_zip_cached(table["url"])
        try:
            raw = zipfile.ZipFile(io.BytesIO(zip_raw)).read(table["zip_member"])
        except KeyError as exc:
            raise ValueError(
                f"El miembro '{table['zip_member']}' ya no está en el ZIP del boletín IEM"
            ) from exc
        try:
            legacy_book = xlrd.open_workbook(file_contents=raw)
        except xlrd.XLRDError as exc:
            raise ValueError("El archivo .xls legado del boletín IEM no se pudo leer") from exc
        worksheet = _XlsSheetAdapter(legacy_book.sheet_by_index(0), legacy_book)
        structured = _extract_wide_series(worksheet, desde, hasta, max_rows)
        if structured is None:
            structured = _extract_long_table(worksheet, desde, hasta, max_rows)
        if structured is None:
            structured = _extract_matrix_series(worksheet, desde, hasta, max_rows)
    elif table.get("formato_origen") == "html_frameset":
        # Pre-Aug-2006 bulletin: the table is a raw HTML <TABLE> on its own
        # section page, not a downloadable file (see
        # _fetch_legacy_frameset_tables). Never attempt structured
        # extraction here -- go straight to the grid preview.
        raw, truncated = await download_bytes(table["url"])
        if truncated:
            raise ValueError("La página de la sección IEM superó el límite de descarga")
        structured = None
    else:
        raw, truncated = await download_bytes(table["url"])
        if truncated:
            raise ValueError("La tabla IEM supera el límite seguro de descarga")
        if not raw.startswith(b"PK"):
            raise ValueError("La URL de la tabla IEM no devolvió un archivo XLSX válido")
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            structured = _extract_wide_series(workbook.active, desde, hasta, max_rows)
            if structured is None:
                structured = _extract_long_table(workbook.active, desde, hasta, max_rows)
            if structured is None:
                structured = _extract_matrix_series(workbook.active, desde, hasta, max_rows)
        finally:
            workbook.close()

    result = {
        "source": _SOURCE_NAME,
        "tabla": {**table, "sha256": hashlib.sha256(raw).hexdigest()},
        "archivo_truncado": False,
    }
    if structured is not None:
        return {**result, **structured}
    if table.get("zip_member"):
        inspector = _inspect_legacy_xls
    elif table.get("formato_origen") == "html_frameset":
        inspector = _inspect_frameset_table
    else:
        inspector = _inspect_xlsx
    return {**result, "formato": "vista", **inspector(raw, max_rows=max_rows)}
