from __future__ import annotations

import hashlib
import io
import zipfile

import openpyxl
import pytest
import xlrd

from helpers import bce_iem_client
from helpers.cache import TtlCache

_INDEX_HTML = """
<a href="m2091062026.html">No. 2091 Mayo 2026</a>
<a href="m2092062026.html">No. 2092 Junio 2026</a>
"""

_ARCHIVE_HTML = """
<a href="/documentos/PublicacionesNotas/Catalogo/IEMensual/Indices/m1727011996.htm">
  <strong>No. 1727 Enero 1996</strong>
</a>
<a href="/documentos/PublicacionesNotas/Catalogo/IEMensual/Indices/m2093072026.html">
  <strong>No. 2093 Julio 2026</strong>
</a>
"""

_BULLETIN_HTML = """
<h2>3. ESTADÍSTICAS DEL SECTOR EXTERNO</h2>
<p>3.1.1 <a href="Catalogo/IEMensual/m2091/IEM-431-e.xlsx">Producto Interno Bruto (PIB): Enfoque del Gasto</a></p>
<p>3.1.2 <a href="Catalogo/IEMensual/m2091/IEM-432-e.xlsx">Exportaciones FOB por Producto Principal</a></p>
<a href="Catalogo/IEMensual/m2091/IEM2091.zip">ZIP completo</a>
"""


_PARSED_BULLETINS = bce_iem_client._parse_bulletins(_INDEX_HTML)
_NEW_BULLETIN_URL = _PARSED_BULLETINS[0]["url"]
_OLD_BULLETIN_URL = _PARSED_BULLETINS[1]["url"]
_PARSED_TABLES = bce_iem_client._parse_tables(_BULLETIN_HTML, _PARSED_BULLETINS[0])
_PIB_TABLE_URL = _PARSED_TABLES[0]["url"]


@pytest.fixture(autouse=True)
def _reset_cache():
    bce_iem_client._catalog_cache = TtlCache(ttl_seconds=60)
    bce_iem_client._bulletins_cache = TtlCache(ttl_seconds=60)
    bce_iem_client._bulletin_tables_cache = TtlCache(ttl_seconds=60, max_entries=512)
    bce_iem_client._bulletin_fetch_locks = {}
    bce_iem_client._zip_bytes_cache = TtlCache(ttl_seconds=60, max_entries=16)
    yield


def _zip_bytes(members: dict[str, bytes]) -> bytes:
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w") as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    return out.getvalue()


class _FakeXlrdCell:
    def __init__(self, ctype: int, value) -> None:
        self.ctype = ctype
        self.value = value


class _FakeXlrdSheet:
    """rows: list of [(ctype, value), ...] -- one tuple per cell."""

    def __init__(self, name: str, rows: list[list[tuple[int, object]]]) -> None:
        self.name = name
        self.nrows = len(rows)
        self.ncols = max((len(row) for row in rows), default=0)
        self._rows = rows

    def cell(self, row_index: int, col_index: int) -> _FakeXlrdCell:
        row = self._rows[row_index]
        if col_index < len(row):
            ctype, value = row[col_index]
        else:
            ctype, value = (xlrd.XL_CELL_EMPTY, "")
        return _FakeXlrdCell(ctype, value)


class _FakeXlrdBook:
    def __init__(self, sheets: list[_FakeXlrdSheet], datemode: int = 0) -> None:
        self.datemode = datemode
        self._sheets = sheets

    def sheet_by_index(self, index: int) -> _FakeXlrdSheet:
        return self._sheets[index]

    def sheets(self) -> list[_FakeXlrdSheet]:
        return self._sheets


def _xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cuadro"
    sheet.append(["Producto Interno Bruto", None, None])
    sheet.append([])
    sheet.append(["Período", 2024, 2025])
    sheet.append(["Variable", "(prev)", "(prev)"])
    sheet.append(["Millones USD", None, None])
    sheet.append(["PIB", 123.4, 130.2])
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _long_xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Nombre de la tabla"])
    sheet.append([])
    sheet.append(["Año", "Indicador", "Valor"])
    sheet.append([2024, "PIB", 100])
    sheet.append([2025, "PIB", 110])
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _monthly_xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Indicador mensual", None, None])
    sheet.append(["Mes", "Ene 2025", "Febrero 2025"])
    sheet.append(["Variable", "valor", "valor"])
    sheet.append(["PIB", 101, 102])
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _wide_xlsx_with_gap() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Período", 2024, 2025])
    sheet.append(["Variable", "(prev)", "(prev)"])
    sheet.append(["Millones USD", None, None])
    sheet.append(["PIB", 123.4, 130.2])
    sheet.append(["PIB Nuevo", 50.0, None])
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def test_parse_bulletins_picks_newest_and_resolves_urls():
    bulletins = bce_iem_client._parse_bulletins(_INDEX_HTML)

    assert [item["numero"] for item in bulletins] == [2092, 2091]
    assert bulletins[0]["url"].endswith("m2092062026.html")


def test_parse_tables_keeps_real_link_not_guessed_bulletin_directory():
    bulletin = bce_iem_client._parse_bulletins(_INDEX_HTML)[0]
    tables = bce_iem_client._parse_tables(_BULLETIN_HTML, bulletin)

    assert [table["table_id"] for table in tables] == ["iem-431-e", "iem-432-e"]
    assert tables[0]["url"].endswith("IEMensual/m2091/IEM-431-e.xlsx")
    assert tables[0]["boletin_numero"] == 2092


def test_merge_bulletins_prefers_latest_publication_for_duplicate_number():
    index = [{"numero": 2092, "url": "index-version"}]
    latest = [{"numero": 2092, "url": "latest-version"}, {"numero": 2093, "url": "new"}]

    merged = bce_iem_client._merge_bulletins(latest, index)

    assert [item["numero"] for item in merged] == [2093, 2092]
    assert merged[1]["url"] == "latest-version"


def test_parse_complete_files_catalogs_pdf_and_zip():
    bulletin = _PARSED_BULLETINS[0]

    files = bce_iem_client._parse_complete_files(
        '<a href="IEM2092.pdf">PDF completo</a>'
        '<a href="IEM2092.zip">ZIP completo</a>'
        '<a href="IEM-431-e.xlsx">Tabla</a>',
        bulletin,
    )

    assert [(item["tipo"], item["nombre"]) for item in files] == [
        ("pdf", "IEM2092.pdf"),
        ("zip", "IEM2092.zip"),
    ]


@pytest.mark.asyncio
async def test_fetch_bulletins_reconciles_latest_publications(httpx_mock):
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(
        url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL,
        html='<a href="m2093072026.html">No. 2093 Julio 2026</a>',
    )
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html=_ARCHIVE_HTML)

    bulletins = await bce_iem_client._fetch_bulletins()

    assert bulletins[0]["numero"] == 2093
    assert len(bulletins) == 4


@pytest.mark.asyncio
async def test_fetch_bulletins_adds_official_historical_archive(httpx_mock):
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL, html="")
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html=_ARCHIVE_HTML)

    bulletins = await bce_iem_client._fetch_bulletins()

    assert [item["numero"] for item in bulletins] == [2093, 2092, 2091, 1727]
    assert bulletins[-1]["anio"] == 1996


@pytest.mark.asyncio
async def test_search_tables_can_persist_the_complete_catalog(httpx_mock, tmp_path, monkeypatch):
    monkeypatch.setenv("IEM_CATALOG_DIR", str(tmp_path))
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(
        url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL, html=""
    )
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html="")
    httpx_mock.add_response(url=_NEW_BULLETIN_URL, html=_BULLETIN_HTML)

    result = await bce_iem_client.search_tables(guardar_catalogo=True)

    assert result["catalogo_guardado"]["archivo"].endswith(".json")
    assert (tmp_path / "latest.json").exists()


def test_merge_historical_tables_keeps_all_versions():
    bulletin_1 = {"numero": 2092, "mes": 6, "anio": 2026, "url": "new"}
    bulletin_2 = {"numero": 2091, "mes": 5, "anio": 2026, "url": "old"}
    tables = bce_iem_client._parse_tables(_BULLETIN_HTML, bulletin_1)
    tables += bce_iem_client._parse_tables(_BULLETIN_HTML, bulletin_2)

    merged = bce_iem_client._merge_historical_tables(tables)

    assert len(merged) == 2
    assert merged[0]["boletines_disponibles"] == 2
    assert [version["boletin_numero"] for version in merged[0]["versiones"]] == [
        2092,
        2091,
    ]


def test_extract_wide_series_returns_none_when_range_not_covered():
    workbook = openpyxl.load_workbook(io.BytesIO(_xlsx()), read_only=True, data_only=True)
    result = bce_iem_client._extract_wide_series(workbook.active, "1990", "1990", 20)
    workbook.close()

    assert result is None


def test_extract_wide_series_keeps_row_with_no_data_in_selected_range():
    workbook = openpyxl.load_workbook(
        io.BytesIO(_wide_xlsx_with_gap()), read_only=True, data_only=True
    )
    result = bce_iem_client._extract_wide_series(workbook.active, "2025", "2025", 20)
    workbook.close()

    assert result is not None
    names = [series["nombre"] for series in result["bloques"][0]["series"]]
    assert names == ["PIB", "PIB Nuevo"]


@pytest.mark.asyncio
async def test_fetch_historical_tables_caps_unbounded_fanout(monkeypatch):
    bulletins = [
        {"numero": n, "mes": 1, "anio": 2000 + n, "url": f"http://x/{n}"}
        for n in range(bce_iem_client._MAX_HISTORICAL_BULLETINS + 20)
    ]

    async def fake_fetch(bulletin: dict) -> list[dict]:
        return [{"table_id": f"t{bulletin['numero']}"}]

    monkeypatch.setattr(bce_iem_client, "_fetch_tables_for_bulletin", fake_fetch)

    _, selected, loaded = await bce_iem_client._fetch_historical_tables(bulletins, 0, 0)

    assert len(selected) == bce_iem_client._MAX_HISTORICAL_BULLETINS
    assert loaded == bce_iem_client._MAX_HISTORICAL_BULLETINS


@pytest.mark.asyncio
async def test_fetch_historical_tables_respects_explicit_range_beyond_cap(monkeypatch):
    total = bce_iem_client._MAX_HISTORICAL_BULLETINS + 20
    bulletins = [
        {"numero": n, "mes": 1, "anio": 2000 + n, "url": f"http://x/{n}"}
        for n in range(total)
    ]

    async def fake_fetch(bulletin: dict) -> list[dict]:
        return [{"table_id": f"t{bulletin['numero']}"}]

    monkeypatch.setattr(bce_iem_client, "_fetch_tables_for_bulletin", fake_fetch)

    _, selected, loaded = await bce_iem_client._fetch_historical_tables(
        bulletins, 2000, 2000 + total - 1
    )

    assert len(selected) == total
    assert loaded == total


def test_extract_long_table_filters_rows_by_year():
    workbook = openpyxl.load_workbook(io.BytesIO(_long_xlsx()), read_only=True, data_only=True)
    result = bce_iem_client._extract_long_table(workbook.active, "2025", "2025", 20)
    workbook.close()

    assert result == {
        "formato": "tabla_larga",
        "encabezados": ["Año", "Indicador", "Valor"],
        "filas": [["2025", "PIB", "110"]],
        "filas_totales": 1,
        "truncada": False,
    }


def test_extract_monthly_wide_series_accepts_spanish_month_labels():
    workbook = openpyxl.load_workbook(
        io.BytesIO(_monthly_xlsx()), read_only=True, data_only=True
    )
    result = bce_iem_client._extract_wide_series(
        workbook.active, "2025-02", "2025-02", 20
    )
    workbook.close()

    assert result is not None
    assert result["periodos"] == ["Febrero 2025"]
    assert result["bloques"][0]["series"] == [
        {"nombre": "PIB", "valores": {"Febrero 2025": 102}}
    ]


def test_period_key_accepts_numeric_month_year_and_spanish_month():
    assert bce_iem_client._period_key("03/2025") == (2025, 3)
    assert bce_iem_client._period_key("Marzo 2025") == (2025, 3)
    assert bce_iem_client._period_key("II trimestre 2025") == (2025, 4)


def test_extract_matrix_series_preserves_descriptor_columns():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Código", "Descripción", 2024, 2025])
    sheet.append(["A", "PIB", 100, 110])

    result = bce_iem_client._extract_matrix_series(sheet, "2025", "2025", 20)

    assert result["formato"] == "series_matriz"
    assert result["columnas_descriptivas"] == ["Código", "Descripción"]
    assert result["bloques"][0]["series"] == [
        {"nombre": "A | PIB", "valores": {"2025": 110}}
    ]


@pytest.mark.asyncio
async def test_search_tables_indexes_latest_bulletin_and_filters(httpx_mock):
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(
        url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL, html=""
    )
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html="")
    httpx_mock.add_response(url=_NEW_BULLETIN_URL, html=_BULLETIN_HTML)

    result = await bce_iem_client.search_tables("exportaciones")

    assert result["boletin"]["numero"] == 2092
    assert result["total"] == 1
    assert result["tablas"][0]["table_id"] == "iem-432-e"


@pytest.mark.asyncio
async def test_search_tables_historical_merges_versions(httpx_mock):
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(
        url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL, html=""
    )
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html="")
    httpx_mock.add_response(url=_NEW_BULLETIN_URL, html=_BULLETIN_HTML)
    httpx_mock.add_response(url=_OLD_BULLETIN_URL, html=_BULLETIN_HTML)

    result = await bce_iem_client.search_tables("PIB", historico=True)

    assert result["historico"] is True
    assert result["boletines_consultados"] == 2
    assert result["boletines_sin_tablas"] == 0
    assert result["tablas"][0]["boletines_disponibles"] == 2


@pytest.mark.asyncio
async def test_get_table_returns_layout_preview(httpx_mock):
    xlsx = _xlsx()
    httpx_mock.add_response(url=bce_iem_client.IEM_INDEX_URL, html=_INDEX_HTML)
    httpx_mock.add_response(
        url=bce_iem_client.IEM_LATEST_PUBLICATIONS_URL, html=""
    )
    httpx_mock.add_response(url=bce_iem_client.IEM_ARCHIVE_URL, html="")
    httpx_mock.add_response(url=_NEW_BULLETIN_URL, html=_BULLETIN_HTML)
    httpx_mock.add_response(url=_PIB_TABLE_URL, content=xlsx)

    result = await bce_iem_client.get_table("iem-431-e", desde="2025", max_rows=3)

    assert result["tabla"]["titulo"].startswith("Producto Interno")
    assert result["formato"] == "series_ancho"
    assert result["periodos"] == ["2025"]
    assert result["bloques"][0]["unidad"] == "Millones USD"
    assert result["bloques"][0]["series"] == [
        {"nombre": "PIB", "valores": {"2025": 130.2}}
    ]
    assert result["tabla"]["sha256"] == hashlib.sha256(xlsx).hexdigest()


# ---- pre-Oct-2016 bulk ZIP fallback (bulletins before No. 1976) -----------


@pytest.mark.parametrize(
    ("filename", "expected"),
    [
        ("IEM-315.xls", "iem-legado-iem-315"),
        ("IEM-315a.xls", "iem-legado-iem-315a"),
        ("5_SectorPetrolero.xls", "iem-legado-5-sectorpetrolero"),
        ("7_GraficosIDEAC.xls", "iem-legado-7-graficosideac"),
    ],
)
def test_legacy_table_id_normalizes_filename(filename, expected):
    assert bce_iem_client._legacy_table_id(filename) == expected


_LEGACY_ZIP_URL = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1975/IEM1975.zip"
_LEGACY_BULLETIN_HTML = f"""
<a href="{_LEGACY_ZIP_URL}">BAJAR PUBLICACIÓN COMPLETA</a>
"""


@pytest.mark.asyncio
async def test_fetch_tables_for_bulletin_falls_back_to_zip_when_no_individual_xlsx(
    httpx_mock,
):
    # No IEM-*-e.xlsx anchors on the page at all -- only the bulk ZIP link,
    # matching every real bulletin before No. 1976 (see module docstring).
    bulletin = dict(_PARSED_BULLETINS[0])
    httpx_mock.add_response(url=bulletin["url"], html=_LEGACY_BULLETIN_HTML)
    zip_url = _LEGACY_ZIP_URL
    httpx_mock.add_response(
        url=zip_url,
        content=_zip_bytes({"IEM-315.xls": b"fake", "5_SectorPetrolero.xls": b"fake"}),
    )

    tables = await bce_iem_client._fetch_tables_for_bulletin(bulletin)

    by_id = {t["table_id"]: t for t in tables}
    assert set(by_id) == {"iem-legado-iem-315", "iem-legado-5-sectorpetrolero"}
    assert by_id["iem-legado-iem-315"]["zip_member"] == "IEM-315.xls"
    assert by_id["iem-legado-iem-315"]["formato_origen"] == "xls_legado_zip"
    assert by_id["iem-legado-iem-315"]["url"] == zip_url


@pytest.mark.asyncio
async def test_fetch_tables_for_bulletin_still_raises_without_zip_or_tables(httpx_mock):
    # A bulletin page with neither individual XLSX links nor a ZIP -- the
    # 4 genuinely dead/incomplete bulletins found in 1996 look like this.
    bulletin = dict(_PARSED_BULLETINS[0])
    httpx_mock.add_response(url=bulletin["url"], html="<p>nada aquí</p>")

    with pytest.raises(ValueError, match="no expuso tablas XLSX individuales"):
        await bce_iem_client._fetch_tables_for_bulletin(bulletin)


@pytest.mark.asyncio
async def test_get_table_reads_legacy_xls_from_zip_member(httpx_mock, monkeypatch):
    bulletin = dict(_PARSED_BULLETINS[0])
    bulletin["numero"] = 1975
    # get_table(boletin_numero=...) resolves via _fetch_bulletins() first --
    # bulletin discovery itself is covered elsewhere, so seed the cache
    # directly instead of also mocking the index/latest/archive pages.
    bce_iem_client._bulletins_cache.set("bulletins", [bulletin])
    httpx_mock.add_response(url=bulletin["url"], html=_LEGACY_BULLETIN_HTML)
    zip_url = _LEGACY_ZIP_URL
    member_bytes = b"not a real xls -- xlrd.open_workbook is mocked below"
    httpx_mock.add_response(
        url=zip_url, content=_zip_bytes({"IEM-315.xls": member_bytes})
    )

    fake_sheet = _FakeXlrdSheet(
        "Cuadro",
        [
            [
                (xlrd.XL_CELL_TEXT, "Período"),
                (xlrd.XL_CELL_NUMBER, 2024.0),
                (xlrd.XL_CELL_NUMBER, 2025.0),
            ],
            [
                (xlrd.XL_CELL_TEXT, "Millones USD"),
                (xlrd.XL_CELL_EMPTY, ""),
                (xlrd.XL_CELL_EMPTY, ""),
            ],
            [
                (xlrd.XL_CELL_TEXT, "PIB"),
                (xlrd.XL_CELL_NUMBER, 123.4),
                (xlrd.XL_CELL_NUMBER, 130.2),
            ],
        ],
    )
    fake_book = _FakeXlrdBook([fake_sheet])
    seen_bytes = []

    def fake_open_workbook(file_contents: bytes) -> _FakeXlrdBook:
        seen_bytes.append(file_contents)
        return fake_book

    monkeypatch.setattr(xlrd, "open_workbook", fake_open_workbook)

    result = await bce_iem_client.get_table(
        "iem-legado-iem-315", desde="2025", boletin_numero=1975
    )

    assert seen_bytes == [member_bytes]
    assert result["tabla"]["formato_origen"] == "xls_legado_zip"
    assert result["tabla"]["sha256"] == hashlib.sha256(member_bytes).hexdigest()
    assert result["formato"] == "series_ancho"
    assert result["bloques"][0]["unidad"] == "Millones USD"
    assert result["bloques"][0]["series"] == [
        {"nombre": "PIB", "valores": {"2025": 130.2}}
    ]


@pytest.mark.asyncio
async def test_download_zip_cached_only_fetches_once(httpx_mock):
    zip_url = _LEGACY_ZIP_URL
    httpx_mock.add_response(
        url=zip_url, content=_zip_bytes({"IEM-315.xls": b"fake"})
    )

    first = await bce_iem_client._download_zip_cached(zip_url)
    second = await bce_iem_client._download_zip_cached(zip_url)

    assert first == second
    assert len(httpx_mock.get_requests()) == 1


# ---- pre-Aug-2006 frameset HTML fallback (bulletins before No. 1854) ------

# Real shape confirmed live on bulletin No. 1800/1780: uppercase, unquoted
# <A HREF = ...>, no closing </TR>/</TH>/</TD>, real ROWSPAN/COLSPAN.
_FRAMESET_BULLETIN_HTML = """
<HTML>
<A HREF = /documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_77.htm TARGET="_top"><IMG SRC = x.gif> <B>1.1 Principales Indicadores</B></A><BR>
<A HREF = /documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_78.htm TARGET="_top"><IMG SRC = x.gif> <B>1.2 Otra Seccion</B></A><BR>
</HTML>
"""

_FRAMESET_SECTION_HTML = """
<HTML>
<HEAD><TITLE> Boletin</TITLE></HEAD>
<BODY>
<H3><CENTER>1.1 PRINCIPALES INDICADORES MONETARIOS</CENTER></H3>
<P>
<TABLE BORDER=2>
<TR>
<TH ROWSPAN=2 colspan=1><B>PERIODO</B></TH>
<TH COLSPAN=2><B>BANCO CENTRAL</B></TH>
</TR>
<TR>
<TH><B>Total</B></TH>
<TH><B>Otro</B></TH>
</TR>
<TR><TD>1999<TD>872.7<TD>577.9
</TABLE>
<HR>
FUENTE: BCE.
</BODY>
</HTML>
"""


def test_table_grid_parser_caps_prose_cell_length():
    # Some "sections" are methodology notes, not data -- one <TD colspan=N>
    # wrapping several paragraphs (confirmed live, bulletin No. 1820), not
    # a parser bug. Must not blow up an otherwise agent-sized response.
    long_html = f"<TABLE><TR><TD colspan=30>{'x' * 2000}</TABLE>"
    parser = bce_iem_client._TableGridParser()
    parser.feed(long_html)

    grid = bce_iem_client._expand_table_grid(parser.tables[0])
    assert len(grid) == 1
    for cell in grid[0]:
        assert len(cell) == bce_iem_client._MAX_FRAMESET_CELL_CHARS + 1  # +1 for "…"
        assert cell.endswith("…")


def test_table_grid_parser_resolves_rowspan_and_colspan():
    parser = bce_iem_client._TableGridParser()
    parser.feed(_FRAMESET_SECTION_HTML)

    assert len(parser.tables) == 1
    grid = bce_iem_client._expand_table_grid(parser.tables[0])
    assert grid == [
        ["PERIODO", "BANCO CENTRAL", "BANCO CENTRAL"],
        ["PERIODO", "Total", "Otro"],
        ["1999", "872.7", "577.9"],
    ]


@pytest.mark.parametrize(
    ("title", "expected"),
    [
        ("1.1 Principales Indicadores Monetarios", "iem-legado-frameset-1-1-principales-indicadores-monetarios"),
        ("", "iem-legado-frameset-seccion-3"),
    ],
)
def test_legacy_frameset_table_id_normalizes_title(title, expected):
    assert bce_iem_client._legacy_frameset_table_id(title, fallback_index=3) == expected


@pytest.mark.asyncio
async def test_fetch_legacy_frameset_tables_discovers_sections(httpx_mock):
    bulletin = dict(_PARSED_BULLETINS[0])
    section1 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_77.htm"
    section2 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_78.htm"
    httpx_mock.add_response(url=section1, content=_FRAMESET_SECTION_HTML.encode("cp1252"))
    httpx_mock.add_response(
        url=section2, content="<H3>1.2 Otra Seccion</H3><TABLE><TR><TD>x</TABLE>".encode("cp1252")
    )

    tables = await bce_iem_client._fetch_legacy_frameset_tables(
        bulletin, _FRAMESET_BULLETIN_HTML
    )

    assert [t["table_id"] for t in tables] == [
        "iem-legado-frameset-1-1-principales-indicadores-monetarios",
        "iem-legado-frameset-1-2-otra-seccion",
    ]
    assert tables[0]["formato_origen"] == "html_frameset"
    assert tables[0]["url"] == section1


@pytest.mark.asyncio
async def test_fetch_tables_for_bulletin_falls_back_to_frameset_when_no_zip(httpx_mock):
    bulletin = dict(_PARSED_BULLETINS[0])
    httpx_mock.add_response(url=bulletin["url"], html=_FRAMESET_BULLETIN_HTML)
    section1 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_77.htm"
    section2 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_78.htm"
    httpx_mock.add_response(url=section1, content=_FRAMESET_SECTION_HTML.encode("cp1252"))
    httpx_mock.add_response(
        url=section2, content="<H3>1.2 Otra Seccion</H3><TABLE><TR><TD>x</TABLE>".encode("cp1252")
    )

    tables = await bce_iem_client._fetch_tables_for_bulletin(bulletin)

    assert len(tables) == 2
    assert tables[0]["formato_origen"] == "html_frameset"


@pytest.mark.asyncio
async def test_get_table_returns_grid_preview_for_frameset_table(httpx_mock):
    bulletin = dict(_PARSED_BULLETINS[0])
    bulletin["numero"] = 1800
    bce_iem_client._bulletins_cache.set("bulletins", [bulletin])
    httpx_mock.add_response(url=bulletin["url"], html=_FRAMESET_BULLETIN_HTML)
    section1 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_77.htm"
    section2 = "https://contenido.bce.fin.ec/documentos/PublicacionesNotas/Catalogo/IEMensual/m1800/m1800_78.htm"
    section_bytes = _FRAMESET_SECTION_HTML.encode("cp1252")
    httpx_mock.add_response(url=section1, content=section_bytes)
    httpx_mock.add_response(
        url=section2, content="<H3>1.2 Otra Seccion</H3><TABLE><TR><TD>x</TABLE>".encode("cp1252")
    )
    # get_table re-downloads the section page to read it -- register the
    # response a second time.
    httpx_mock.add_response(url=section1, content=section_bytes)

    result = await bce_iem_client.get_table(
        "iem-legado-frameset-1-1-principales-indicadores-monetarios",
        boletin_numero=1800,
    )

    assert result["formato"] == "vista"
    assert result["tabla"]["formato_origen"] == "html_frameset"
    assert result["hojas"][0]["vista"] == [
        ["PERIODO", "BANCO CENTRAL", "BANCO CENTRAL"],
        ["PERIODO", "Total", "Otro"],
        ["1999", "872.7", "577.9"],
    ]
    assert result["tabla"]["sha256"] == hashlib.sha256(section_bytes).hexdigest()
